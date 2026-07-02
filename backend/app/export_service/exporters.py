"""
Downloads module: CSV / Excel / PDF shortlist export. Each function takes
the same list of ranking-result dicts (the API's RankingResultOut shape)
and returns raw bytes, so the route layer just sets the right
content-type/filename and streams the response.
"""
from __future__ import annotations

import io
from typing import Any, Dict, List

EXPORT_COLUMNS = [
    "Rank", "Candidate Name", "Overall Score", "Confidence",
    "Risk", "Recommendation", "Missing Skills", "Reasoning",
]


def _row_for(r: Dict[str, Any]) -> List[Any]:
    return [
        r.get("rank_position"),
        r.get("candidate_name"),
        r["scores"]["overall_score"],
        f"{r.get('confidence_score')}% ({r.get('confidence_band')})",
        r.get("risk_level"),
        r.get("recommendation"),
        ", ".join(r.get("missing_skills", [])),
        (r.get("reasoning_text") or "")[:300],
    ]


def export_csv(results: List[Dict[str, Any]]) -> bytes:
    import csv

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(EXPORT_COLUMNS)
    for r in results:
        writer.writerow(_row_for(r))
    return buffer.getvalue().encode("utf-8")


def export_excel(results: List[Dict[str, Any]], jd_title: str = "") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Shortlist"

    ws.append([f"AI Recruiter Platform -- Shortlist for: {jd_title}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(EXPORT_COLUMNS))
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="7C3AED")
    ws.append([])

    header_row = 3
    ws.append(EXPORT_COLUMNS)
    for col_idx in range(1, len(EXPORT_COLUMNS) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")
        cell.alignment = Alignment(horizontal="center")

    risk_fill = {
        "Low": PatternFill("solid", fgColor="D1FAE5"),
        "Medium": PatternFill("solid", fgColor="FEF3C7"),
        "High": PatternFill("solid", fgColor="FEE2E2"),
    }

    for r in results:
        row = _row_for(r)
        ws.append(row)
        risk_cell = ws.cell(row=ws.max_row, column=5)
        if r.get("risk_level") in risk_fill:
            risk_cell.fill = risk_fill[r["risk_level"]]
        ws.cell(row=ws.max_row, column=8).alignment = Alignment(wrap_text=True, vertical="top")

    widths = [6, 22, 14, 18, 10, 24, 30, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_pdf(results: List[Dict[str, Any]], jd_title: str = "") -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), title="Candidate Shortlist")
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"AI Recruiter Platform -- Shortlist Report", styles["Title"]),
        Paragraph(f"Job: {jd_title}", styles["Heading3"]),
        Spacer(1, 12),
    ]

    table_data = [EXPORT_COLUMNS]
    for r in results:
        row = _row_for(r)
        row[7] = Paragraph(str(row[7]), styles["BodyText"])
        table_data.append(row)

    table = Table(table_data, repeatRows=1, colWidths=[35, 90, 55, 70, 45, 90, 110, 220])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
    ]))
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()
